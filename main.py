from selenium import webdriver
import openpyxl as o
import time

#These will need to point to the files as they are located on your system. Note the '/' NOT the '\' being used
#Workbook Copy needs to be just that- a copy of the workbook.
WORKBOOK_LOCATION = "C:/Users/kyleb/Documents/Programs/AdamProject/Listings.xlsx"
WORKBOOK_COPY_LOCATION = "C:/Users/kyleb/Documents/Programs/AdamProject/ListingsCopy.xlsx"
SAVE_FILE_NAME = "ListingsCopy.xlsx"

#To run with chrome, had to download chromedriver
#http://chromedriver.chromium.org/downloads
#Once downloaded, unzip and place chromedriver.exe in desired location
WEB_DRIVER_LOCATION = "C:/Users/kyleb/Documents/Programs/AdamProject/chromedriver.exe"
#WEBPAGE_URL = "http://qpublic5.qpublic.net/psp/sc_pickens_parcel.php"

#These values represent the spreadsheet row start and end values.
#Note the first row in excel is indexed as ONE, not ZERO
ROW_START = 2
ROW_END = 582

driver = webdriver.Chrome(WEB_DRIVER_LOCATION)

def main():

    parcelNumbersList = getParcelNumbersFromWorksheet()

    propertyAddressList, accountNumberList = getPropertyAddresses(parcelNumbersList)

    ownerAddressList = getOwnerAddress(parcelNumbersList, accountNumberList)
 
    saveToWorkbook(propertyAddressList, ownerAddressList)

    driver.quit()

def getParcelNumbersFromWorksheet():
    parcelNums = []
    
    workBook = o.load_workbook(WORKBOOK_LOCATION)
    sheet = workBook.active

    for row in sheet.iter_rows(min_row=ROW_START, max_row=ROW_END):
        parcelNums.append(row[1].value)

    workBook.close()
    
    return parcelNums

def getPropertyAddresses(parcelNumbersList):

    firstPartURL = "http://qpublic5.qpublic.net/qp5/sc_pickens_alsearch.php?desc=false&Parcel_Search=Search+By+Parcel+ID&searchType=parcel_id&INPUT="
    secondPartURL = "&BEGIN=0&order=parcel"

    propertyAddressList = []
    accountNumberList = []

    for number in parcelNumbersList:
        driver.get(firstPartURL + number + secondPartURL)
        try:
            address = driver.find_elements_by_class_name("search_value")[3].text
            account = driver.find_elements_by_class_name("search_value")[1].text.strip()
            if (address == None or address.strip() == ""):
                address = 'ADDRESS NOT FOUND'
            if (account == None or account.strip() == ""):
                account = 'ACCOUNT NOT FOUND'
                
            propertyAddressList.append(address)
            accountNumberList.append(account)
        except:
            address = 'NOT FOUND BAD INPUT'
            propertyAddressList.append(address)
            account = 'NOT FOUND BAD INPUT'
            accountNumberList.append(account)
        
    
    return propertyAddressList, accountNumberList

def getOwnerAddress(parcelNumbersList, accountNumberList):
    
    firstPartURL = "http://qpublic5.qpublic.net/qp5/sc_pickens_display.php?account="
    secondPartURL = "&KEY="

    ownerAddressList = []
    
    for number in parcelNumbersList:
        driver.get(firstPartURL + accountNumberList[parcelNumbersList.index(number)] + secondPartURL + number)
        try:
            ownerAddress = driver.find_elements_by_class_name("owner_value")[2].text + driver.find_elements_by_class_name("owner_value")[4].text
            if (ownerAddress == None or ownerAddress.strip() == ""):
                ownerAddress = 'NOT FOUND'

            ownerAddressList.append(ownerAddress)

        except:
            ownerAddress = 'NOT FOUND BAD INPUT'
            ownerAddressList.append(ownerAddress)

    return ownerAddressList
    

def saveToWorkbook(propertyAddressList, ownerAddressList):
    workBook = o.load_workbook(WORKBOOK_COPY_LOCATION)
    sheet = workBook.active

    index = 0
    for row in sheet.iter_rows(min_row=ROW_START, max_row=ROW_END):
        row[3].value = propertyAddressList[index]
        row[5].value = ownerAddressList[index]
        index = index + 1

    workBook.save(SAVE_FILE_NAME)
    workBook.close()

if __name__ == "__main__":
    main()
